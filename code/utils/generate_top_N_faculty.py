"""
utils/generate_top_N_faculty.py
--------------------------------
Automatically generates the Top-N faculty list for WoS and ScholarGPS extraction.

This script runs automatically after Google Scholar aggregation (or can be run
manually via main.py option [G]).

Workflow
--------
1. Reads Google Scholar citation rankings (last 5 years)
2. Takes the top N faculty by total citations (N = config.TOP50_N, default 50)
3. Looks up each faculty member's WoS ResearchID and ScholarGPS URL from the
   master ID file (2026_Spring_Faculty_List.xlsx)
4. Flags faculty who are missing WoS IDs or ScholarGPS URLs so you know
   who to contact
5. Saves a formatted Excel file ready to use as input for WoS and ScholarGPS
   extractors

Output columns (matches what WoS and ScholarGPS extractors expect)
------
  Last Name, First Name, Department, Email, Faculty Type,
  WoS ResearchID, scholargps, comments

Prerequisites
-------------
- Google Scholar aggregation must have been run first:
    results/google_scholar/Google_Scholar_Citations_Last_Five_Years.csv
- The master ID file must have a 'scholargps' column added:
    2026_Spring_Faculty_List.xlsx  →  add 'scholargps' column with profile URLs

Run via main.py or directly:
    python -m utils.generate_top50

Output
------
data/Top_{N}_Faculty_{year}.xlsx   (e.g. data/Top_50_Faculty_2026.xlsx)
"""

import sys
import pandas as pd
from pathlib import Path

if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent.parent))

import config

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


def run() -> Path | None:
    """
    Main entry point. Returns the output file path on success, None on failure.
    Called by main.py after Google Scholar aggregation, or run directly.
    """
    print(f"\n[Top-{config.TOP50_N} Generator] Building top faculty list…")

    # ── Check GS citations file exists ────────────────────────────────────────
    gs_csv = config.GS_RESULTS / "Google_Scholar_Citations_Last_Five_Years.csv"
    if not gs_csv.exists():
        print(f"  [ERROR] Google Scholar citations not found: {gs_csv}")
        print("  Run Google Scholar extraction first (option [2] in main.py).")
        return None

    # ── Check master ID file exists ───────────────────────────────────────────
    if not config.FACULTY_FULL_PATH.exists():
        print(f"  [ERROR] Master ID file not found: {config.FACULTY_FULL_PATH}")
        print("  Expected: 2026_Spring_Faculty_List.xlsx in the project root.")
        return None

    # ── Load GS citation rankings ─────────────────────────────────────────────
    df_gs = pd.read_csv(gs_csv)
    df_gs[config.COL_TOTAL_CITES] = pd.to_numeric(
        df_gs[config.COL_TOTAL_CITES], errors="coerce"
    ).fillna(0)

    # Take top N by total citations
    df_top = (
        df_gs
        .sort_values(config.COL_TOTAL_CITES, ascending=False)
        .head(config.TOP50_N)
        .reset_index(drop=True)
    )
    print(f"  Selected top {len(df_top)} faculty from Google Scholar rankings")

    # ── Load master ID file ───────────────────────────────────────────────────
    df_ids = pd.read_excel(config.FACULTY_FULL_PATH)

    # Normalise for reliable matching
    def _norm(s: pd.Series) -> pd.Series:
        return s.fillna("").astype(str).str.strip().str.lower()

    df_ids["_email"] = _norm(df_ids["Email"])
    df_ids["_last"]  = _norm(df_ids["Last Name"])
    df_ids["_first"] = _norm(df_ids["First Name"])

    df_top["_email"] = _norm(df_top[config.COL_EMAIL])
    df_top["_last"]  = _norm(df_top[config.COL_LAST_NAME])
    df_top["_first"] = _norm(df_top[config.COL_FIRST_NAME])

    # Deduplicate master list before indexing
    ids_by_email = df_ids.drop_duplicates("_email").set_index("_email")
    ids_by_name  = df_ids.drop_duplicates(["_last", "_first"]).set_index(["_last", "_first"])

    # Determine which ID columns exist in the master file
    wos_col  = "WoS"        if "WoS"        in df_ids.columns else None
    sgps_col = "scholargps" if "scholargps" in df_ids.columns else None

    if wos_col is None:
        print("  [WARN] No 'WoS' column found in master ID file — WoS ResearchID will be blank.")
    if sgps_col is None:
        print("  [WARN] No 'scholargps' column found in master ID file.")
        print("         Add a 'scholargps' column to 2026_Spring_Faculty_List.xlsx with profile URLs.")

    # ── Look up IDs for each top faculty member ───────────────────────────────
    wos_ids    = []
    sgps_urls  = []
    comments   = []

    for _, row in df_top.iterrows():
        email = row["_email"]
        last  = row["_last"]
        first = row["_first"]

        # Try email match first, then name match
        if email and email in ids_by_email.index:
            match = ids_by_email.loc[email]
        elif (last, first) in ids_by_name.index:
            match = ids_by_name.loc[(last, first)]
        else:
            match = None

        # WoS ResearchID
        wos_id = ""
        if match is not None and wos_col and pd.notna(match.get(wos_col, "")):
            wos_id = str(match[wos_col]).strip()
        wos_ids.append(wos_id)

        # ScholarGPS URL
        sgps_url = ""
        if match is not None and sgps_col and pd.notna(match.get(sgps_col, "")):
            sgps_url = str(match[sgps_col]).strip()
        sgps_urls.append(sgps_url)

        # Build comment flags
        flags = []
        if not wos_id:
            flags.append("NEED WoS ID")
        if not sgps_url:
            flags.append("NEED ScholarGPS URL")
        if match is None:
            flags.append("NOT FOUND IN MASTER LIST")
        comments.append(" | ".join(flags))

    # ── Build output DataFrame ────────────────────────────────────────────────
    df_out = pd.DataFrame({
        config.COL_RANK:         range(1, len(df_top) + 1),
        config.COL_LAST_NAME:    df_top[config.COL_LAST_NAME].values,
        config.COL_FIRST_NAME:   df_top[config.COL_FIRST_NAME].values,
        config.COL_DEPARTMENT:   df_top[config.COL_DEPARTMENT].values,
        config.COL_EMAIL:        df_top[config.COL_EMAIL].values,
        config.COL_FACULTY_TYPE: df_top[config.COL_FACULTY_TYPE].values,
        "GS_Total_Citations":    df_top[config.COL_TOTAL_CITES].values,
        "WoS ResearchID":        wos_ids,
        "scholargps":            sgps_urls,
        "comments":              comments,
    })

    # ── Save Excel with formatting ─────────────────────────────────────────────
    config.DATA_DIR.mkdir(parents=True, exist_ok=True)
    out_path = config.FACULTY_TOP50_PATH
    df_out.to_excel(out_path, index=False)

    if OPENPYXL_OK:
        _format_excel(out_path, df_out)

    # ── Print summary ─────────────────────────────────────────────────────────
    need_wos  = sum(1 for c in comments if "NEED WoS ID"        in c)
    need_sgps = sum(1 for c in comments if "NEED ScholarGPS URL" in c)
    not_found = sum(1 for c in comments if "NOT FOUND"           in c)

    print(f"\n[Top-{config.TOP50_N} Generator] Saved → {out_path.name}")
    print(f"  Faculty selected   : {len(df_out)}")
    print(f"  WoS IDs found      : {len(df_out) - need_wos} / {len(df_out)}")
    print(f"  ScholarGPS URLs    : {len(df_out) - need_sgps} / {len(df_out)}")
    if not_found:
        print(f"  ⚠  Not in master list: {not_found} (check spelling / email)")
    if need_wos or need_sgps:
        items = []
        if need_wos:
            items.append("WoS IDs")
        if need_sgps:
            items.append("ScholarGPS URLs")
        print(f"\n  Action required: open {out_path.name} and fill in missing {' and '.join(items)} (highlighted rows).")

    return out_path


def _format_excel(path: Path, df: pd.DataFrame) -> None:
    """Apply colour formatting to the generated Excel file."""
    wb = load_workbook(path)
    ws = wb.active

    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue header
    ALERT_FILL   = PatternFill("solid", fgColor="FFD700")   # gold = needs attention
    OK_FILL      = PatternFill("solid", fgColor="E2EFDA")   # light green = complete
    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    header_vals = [cell.value for cell in ws[1]]
    comments_col_idx = header_vals.index("comments") + 1  # 1-based

    # Header row
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER

    # Data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        comment_val = str(row[comments_col_idx - 1].value or "")
        needs_action = (
            "NEED WoS ID" in comment_val
            or "NEED ScholarGPS URL" in comment_val
            or "NOT FOUND" in comment_val
        )

        for cell in row:
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center")

        if needs_action:
            for cell in row:
                cell.fill = ALERT_FILL
        else:
            for cell in row:
                cell.fill = OK_FILL

    # Column widths
    col_widths = {
        "Rank": 6, "Last Name": 16, "First Name": 14,
        "Department": 12, "Email": 32, "Faculty Type": 12,
        "GS_Total_Citations": 14, "WoS ResearchID": 18,
        "scholargps": 55, "comments": 35,
    }
    for i, col_name in enumerate(header_vals, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    run()
