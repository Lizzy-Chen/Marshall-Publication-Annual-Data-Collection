"""
utils/comparison.py
--------------------
Cross-source comparison and visualisation.

Merges citation rankings from all available sources (WoS, Google Scholar,
ScholarGPS), computes average rank across sources, ranks faculty by that
average rank (lower = better), and produces:
  - Comparison_Ranked.csv                (raw merged data)
  - Citation_Comparison_All_Sources.xlsx (formatted Excel — all sources)
  - comparison_chart_all_sources.png     (grouped bar chart — all sources)

Called by main.py after extraction/aggregation, or run directly:
    python -m utils.comparison

Outputs
-------
results/comparison/
"""

import sys
import pandas as pd
from pathlib import Path

if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent.parent))

import config

try:
    import matplotlib
    matplotlib.use("Agg")   # Non-interactive backend — works without a display
    import matplotlib.pyplot as plt
    import matplotlib.ticker as ticker
    import numpy as np
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("[Comparison] WARNING: matplotlib not installed. Chart will be skipped.")


# ── Source display metadata ────────────────────────────────────────────────────
# Maps internal source key → (rank col, citations col, link col, bar colour, header fill)
SOURCE_META = {
    "google_scholar": (
        "Google_Rank", "Google_Total Citations",
        "Google Scholar Profile Link",
        "#1f77b4",   # bar colour: blue
        "BDD7EE",    # Excel header fill: light blue
    ),
    "wos": (
        "WoS_Rank", "WoS_Total Citations",
        "WoS ResearchID",
        "#d62728",   # bar colour: red
        "FCE4D6",    # Excel header fill: light salmon
    ),
    "scholargps": (
        "ScholarGPS_Rank", "ScholarGPS_Total Citations",
        "ScholarGPS Profile Link",
        "#2ca02c",   # bar colour: green
        "E2EFDA",    # Excel header fill: light green
    ),
}

# Ratio definitions: (output column name, numerator source, denominator source)
# Expected trend: GS > WoS > ScholarGPS.
#   GS_WoS_Ratio:   GS ÷ WoS  — HIGH outlier = WoS under-counting
#   SGPS_GS_Ratio:  SGPS ÷ GS  — LOW  outlier = SGPS under-counting / wrong profile
#   SGPS_WoS_Ratio: SGPS ÷ WoS — LOW  outlier = SGPS under-counting / wrong profile
_RATIO_DEFS = [
    ("GS_WoS_Ratio",   "google_scholar", "wos"),         # high = WoS under-counting
    ("SGPS_GS_Ratio",  "scholargps",     "google_scholar"),  # low = SGPS under-counting
    ("SGPS_WoS_Ratio", "scholargps",     "wos"),             # low = SGPS under-counting
]


def _load_citations_csv(path: Path, source_label: str) -> pd.DataFrame | None:
    """
    Load a per-source citation rankings CSV.
    Returns None (with a warning) if the file doesn't exist.
    """
    if not path.exists():
        print(f"  [SKIP] {source_label} citations file not found: {path.name}")
        return None
    return pd.read_csv(path)


def run() -> None:
    """
    Main comparison entry point. Discovers which source result files exist,
    merges them, and produces comparison outputs.
    """
    print("\n[Comparison] Starting cross-source comparison…")

    config.COMP_RESULTS.mkdir(parents=True, exist_ok=True)

    # ── Load available source rankings ────────────────────────────────────────
    sources_loaded = {}

    wos_path = config.WOS_RESULTS / "WoS_Citations_Last_Five_Years.csv"
    df_wos   = _load_citations_csv(wos_path, "WoS")
    if df_wos is not None:
        df_wos = df_wos.rename(columns={
            config.COL_RANK:        "WoS_Rank",
            config.COL_TOTAL_CITES: "WoS_Total Citations",
        })
        sources_loaded["wos"] = df_wos

    gs_path = config.GS_RESULTS / "Google_Scholar_Citations_Last_Five_Years.csv"
    df_gs   = _load_citations_csv(gs_path, "Google Scholar")
    if df_gs is not None:
        df_gs = df_gs.rename(columns={
            config.COL_RANK:        "Google_Rank",
            config.COL_TOTAL_CITES: "Google_Total Citations",
        })
        sources_loaded["google_scholar"] = df_gs

    sgps_path = config.SGPS_RESULTS / "ScholarGPS_Citations_Last_Five_Years.csv"
    df_sgps   = _load_citations_csv(sgps_path, "ScholarGPS")
    if df_sgps is not None:
        df_sgps = df_sgps.rename(columns={
            config.COL_RANK:        "ScholarGPS_Rank",
            config.COL_TOTAL_CITES: "ScholarGPS_Total Citations",
        })
        sources_loaded["scholargps"] = df_sgps

    if len(sources_loaded) < 2:
        print("[Comparison] Need at least 2 sources to compare. Run more sources first.")
        return

    print(f"  Sources available: {', '.join(sources_loaded.keys())}")

    # ── Merge all sources (INNER JOIN on faculty key columns) ─────────────────
    source_list = list(sources_loaded.values())
    df_merged   = source_list[0]
    for df_next in source_list[1:]:
        df_merged = df_merged.merge(df_next, on=config.FACULTY_KEY_COLS, how="inner")

    # ── Compute average rank across available sources ─────────────────────────
    # Lower average rank = consistently ranked higher across sources.
    rank_cols = [c for c in df_merged.columns if c.endswith("_Rank")]
    df_merged["Average_Rank"] = df_merged[rank_cols].mean(axis=1).round(2)

    df_merged = df_merged.sort_values(
        "Average_Rank", ascending=True
    ).reset_index(drop=True)
    df_merged.insert(0, "Overall_Rank", range(1, len(df_merged) + 1))

    # ── Compute citation ratios ───────────────────────────────────────────────
    cit_map = {src: SOURCE_META[src][1] for src in sources_loaded if src in SOURCE_META}
    ratio_cols_present = []
    for ratio_col, num_src, den_src in _RATIO_DEFS:
        if num_src in cit_map and den_src in cit_map:
            df_merged[ratio_col] = (
                df_merged[cit_map[num_src]]
                / df_merged[cit_map[den_src]].replace(0, float("nan"))
            ).round(2)
            ratio_cols_present.append(ratio_col)

    # ── Reorder columns: rank → key cols → links/IDs → source rank+cit → average → ratios ──
    priority_order = ["Overall_Rank"] + config.FACULTY_KEY_COLS

    # All profile links / IDs grouped together after basic info
    for src_key in ["google_scholar", "wos", "scholargps"]:
        if src_key in SOURCE_META and src_key in sources_loaded:
            _, _, link_col, _, _ = SOURCE_META[src_key]
            if link_col in df_merged.columns:
                priority_order.append(link_col)

    # Per-source rank and citations
    for src_key in ["google_scholar", "wos", "scholargps"]:
        if src_key in SOURCE_META and src_key in sources_loaded:
            rank_col, cit_col, _, _, _ = SOURCE_META[src_key]
            for col in [rank_col, cit_col]:
                if col in df_merged.columns:
                    priority_order.append(col)

    priority_order.append("Average_Rank")
    priority_order.extend(ratio_cols_present)

    remaining  = [c for c in df_merged.columns if c not in priority_order]
    df_merged  = df_merged[priority_order + remaining]

    # ── Save Comparison_Ranked.csv ────────────────────────────────────────────
    ranked_csv = config.COMP_RESULTS / "Comparison_Ranked.csv"
    df_merged.to_csv(ranked_csv, index=False, encoding="utf-8")
    print(f"  Saved: {ranked_csv.name} ({len(df_merged)} faculty)")

    # ── Save unified Excel and chart ──────────────────────────────────────────
    _save_unified_excel(df_merged, sources_loaded)
    _save_comparison_chart(df_merged, sources_loaded)

    print("[Comparison] Complete.")

    # ── Auto-run outlier report ───────────────────────────────────────────────
    from utils import outlier_report
    outlier_report.run()

    # ── Print top 5 by average rank ───────────────────────────────────────────
    print("\n  Top 5 by average rank across sources:")
    for _, r in df_merged.head(5).iterrows():
        print(f"    #{int(r['Overall_Rank'])} "
              f"{r[config.COL_FIRST_NAME]} {r[config.COL_LAST_NAME]} "
              f"— avg rank {r['Average_Rank']:.2f}")


def _save_unified_excel(df_merged: pd.DataFrame, sources_loaded: dict) -> None:
    """
    Save a single formatted Excel file comparing all available sources.

    Layout
    ------
    - One header row, dark navy with white bold text for identity/link columns
    - Source columns use distinct light-tint fills:
        Google Scholar → light blue
        WoS            → light salmon
        ScholarGPS     → light green
        Average        → purple (white text)
        Ratios         → light lavender
    - Ratio cell highlighting (outliers only, not whole rows):
        GS:WoS ratio HIGH outlier  → gold cell   (WoS under-counting)
        SGPS:GS ratio LOW outlier  → orange cell (SGPS under-counting)
        SGPS:WoS ratio LOW outlier → amber cell  (SGPS under-counting)
        Zero WoS + GS ≥ 20        → light red on WoS citation cell
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [SKIP] openpyxl not installed — Excel comparison skipped.")
        return

    out_path = config.COMP_RESULTS / "Citation_Comparison_All_Sources.xlsx"
    df_merged.to_excel(out_path, index=False)

    wb = load_workbook(out_path)
    ws = wb.active

    # ── Colour palette ─────────────────────────────────────────────────────────
    HEADER_BASE  = PatternFill("solid", fgColor="1F4E79")  # dark navy
    HEADER_AVG   = PatternFill("solid", fgColor="7030A0")  # purple
    HEADER_RATIO = PatternFill("solid", fgColor="D9D2E9")  # light lavender
    BORDER = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    # Source-specific header fill colours (light tints)
    SOURCE_HEADER_FILL = {
        src_key: PatternFill("solid", fgColor=meta[4])
        for src_key, meta in SOURCE_META.items()
    }
    # Map column name → source key (for header colouring)
    col_to_source = {}
    for src_key, (rank_col, cit_col, link_col, _, _) in SOURCE_META.items():
        col_to_source[rank_col] = src_key
        col_to_source[cit_col]  = src_key
        col_to_source[link_col] = src_key

    RATIO_COL_NAMES = {"GS_WoS_Ratio", "SGPS_GS_Ratio", "SGPS_WoS_Ratio"}

    header_vals = [cell.value for cell in ws[1]]

    # ── Pre-compute ratio outlier thresholds (SD-based, matching outlier_report.py) ──
    # GS_WoS_Ratio: flag HIGH (mean + SD_MULTIPLIER × SD) — WoS under-counting.
    # SGPS ratios:  flag LOW  (mean − SD_MULTIPLIER × SD) — SGPS under-counting.
    from utils.outlier_report import GS_WOS_LO, GS_WOS_HI, SD_MULTIPLIER, MIN_GS_FOR_ZERO_WOS_FLAG
    ratio_hi_thresholds = {}   # rc → upper threshold (flag when value > this)
    ratio_lo_thresholds = {}   # rc → lower threshold (flag when value < this)

    # GS:WoS uses fixed bounds from professor's domain knowledge (< 1.5 or > 3.5)
    ratio_lo_thresholds["GS_WoS_Ratio"] = GS_WOS_LO
    ratio_hi_thresholds["GS_WoS_Ratio"] = GS_WOS_HI

    for rc in ["SGPS_GS_Ratio", "SGPS_WoS_Ratio"]:
        if rc in df_merged.columns:
            vals = df_merged[rc].dropna()
            if len(vals) >= 4:
                ratio_lo_thresholds[rc] = float(vals.mean() - SD_MULTIPLIER * vals.std())

    wos_cit_col = SOURCE_META["wos"][1] if "wos" in sources_loaded else None
    gs_cit_col  = SOURCE_META["google_scholar"][1] if "google_scholar" in sources_loaded else None

    # Ratio outlier fills (applied to the specific ratio cell, not the whole row)
    RATIO_FILLS = {
        "GS_WoS_Ratio":   PatternFill("solid", fgColor="FFD700"),  # gold  — high GS:WoS (WoS under-counting)
        "SGPS_GS_Ratio":  PatternFill("solid", fgColor="F4B942"),  # orange — low SGPS:GS
        "SGPS_WoS_Ratio": PatternFill("solid", fgColor="FFC000"),  # amber  — low SGPS:WoS
    }
    ZERO_WOS_FILL = PatternFill("solid", fgColor="FFCCCC")  # light red

    # ── Header row ─────────────────────────────────────────────────────────────
    for cell in ws[1]:
        col_name = cell.value
        src_key  = col_to_source.get(col_name)
        if col_name == "Average_Rank":
            fill, font_color = HEADER_AVG, "FFFFFF"
        elif col_name in RATIO_COL_NAMES:
            fill, font_color = HEADER_RATIO, "000000"
        elif src_key and src_key in SOURCE_HEADER_FILL:
            fill, font_color = SOURCE_HEADER_FILL[src_key], "000000"
        else:
            fill, font_color = HEADER_BASE, "FFFFFF"

        cell.fill      = fill
        cell.font      = Font(bold=True, color=font_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row_dict = {header_vals[i]: row[i].value for i in range(len(header_vals))}

        # Determine per-cell fills
        cell_fills = {}

        # Zero WoS + significant GS → flag WoS citation cell in light red
        if wos_cit_col and gs_cit_col:
            wos_v = row_dict.get(wos_cit_col) or 0
            gs_v  = row_dict.get(gs_cit_col) or 0
            if wos_v == 0 and gs_v >= MIN_GS_FOR_ZERO_WOS_FLAG:
                cell_fills[wos_cit_col] = ZERO_WOS_FILL

        # Ratio outlier → highlight the specific ratio cell
        for rc, fill in RATIO_FILLS.items():
            val = row_dict.get(rc)
            if val is None:
                continue
            if rc in ratio_hi_thresholds and val > ratio_hi_thresholds[rc]:
                cell_fills[rc] = fill  # GS:WoS too high → WoS under-counting
            elif rc in ratio_lo_thresholds and val < ratio_lo_thresholds[rc]:
                cell_fills[rc] = fill  # SGPS ratio too low → SGPS under-counting

        for cell in row:
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center")
            col_name = header_vals[cell.column - 1]
            if col_name in cell_fills:
                cell.fill = cell_fills[col_name]
            if col_name in RATIO_COL_NAMES or col_name == "Average_Rank":
                cell.number_format = "0.00"

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = {
        "Overall_Rank": 8,
        "Last Name": 16,  "First Name": 14, "Department": 12,
        "Email": 30,      "Faculty Type": 12,
        "Google Scholar Profile Link": 55,
        "WoS ResearchID": 18,
        "ScholarGPS Profile Link": 55,
        "Google_Rank": 10, "Google_Total Citations": 14,
        "WoS_Rank": 10,    "WoS_Total Citations": 14,
        "ScholarGPS_Rank": 10, "ScholarGPS_Total Citations": 14,
        "Average_Rank": 12,
        "GS_WoS_Ratio": 12, "SGPS_GS_Ratio": 12, "SGPS_WoS_Ratio": 12,
    }
    for i, col_name in enumerate(header_vals, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 12)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    wb.save(out_path)
    print(f"  Excel comparison saved → {out_path.name}")


def _save_comparison_chart(df_merged: pd.DataFrame, sources_loaded: dict) -> None:
    """
    Save a single grouped bar chart showing citations per source for the top 30
    faculty, sorted by average citation rank.
    """
    if not MATPLOTLIB_AVAILABLE:
        return

    # Collect available source series
    series = []
    for src_key, (_, cit_col, _link, colour, _fill) in SOURCE_META.items():
        if src_key in sources_loaded and cit_col in df_merged.columns:
            label = {"google_scholar": "Google Scholar",
                     "wos": "Web of Science",
                     "scholargps": "ScholarGPS"}[src_key]
            series.append((label, cit_col, colour))

    if not series:
        return

    df_plot = df_merged.sort_values("Overall_Rank").head(30)
    names   = df_plot[config.COL_LAST_NAME] + ", " + df_plot[config.COL_FIRST_NAME]
    x       = np.arange(len(names))
    n       = len(series)
    width   = 0.8 / n   # bars side by side within each faculty slot

    fig, ax = plt.subplots(figsize=(16, 7))

    for i, (label, cit_col, colour) in enumerate(series):
        offsets = x + (i - (n - 1) / 2) * width
        ax.bar(offsets, df_plot[cit_col], width=width,
               label=label, color=colour, alpha=0.85)

    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Total Citations (Last 5 Years)")
    ax.set_title(
        f"Citation Comparison: {', '.join(s[0] for s in series)}\n"
        f"(Top {len(df_plot)} faculty by average rank — {config.YEAR_START}–{config.YEAR_END})"
    )
    ax.legend()
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{int(v):,}"))
    ax.grid(axis="y", linestyle="--", alpha=0.4)

    plt.tight_layout()
    out_path = config.COMP_RESULTS / "comparison_chart_all_sources.png"
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Chart saved → {out_path.name}")


if __name__ == "__main__":
    run()
