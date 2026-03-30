"""
utils/outlier_report.py
-----------------------
Generates a flagged Excel report identifying citation anomalies across data sources.

Expected citation trend: Google Scholar > WoS > ScholarGPS.

Three outlier categories are detected:

  1. Zero WoS citations (with significant GS citations)
     Faculty with zero WoS citations but GS citations ≥ MIN_GS_FOR_ZERO_WOS_FLAG
     are flagged — their WoS ResearchID is likely missing, wrong, or stale.

  2. Abnormal GS:WoS ratio (either direction)
     GS:WoS = Google Scholar ÷ WoS citations. Typical cohort value is ~2.5 (SD ~1–1.5).
     Faculty whose ratio falls outside mean ± SD_MULTIPLIER × cohort SD are flagged:
       High ratio (> mean + SD_MULTIPLIER×SD): WoS under-counting — verify ResearchID.
       Low ratio  (< mean − SD_MULTIPLIER×SD): GS over-counting — verify GS profile.
     Expects ~12 flagged faculty in a cohort of ~50.

  3. Low SGPS:GS ratio (ScholarGPS under-counting)
     SGPS:GS = ScholarGPS ÷ GS citations. Expected to be < 1 (SGPS < GS).
     Faculty whose ratio falls below mean − SD_MULTIPLIER × cohort SD are flagged —
     ScholarGPS citations are unexpectedly low, likely due to a wrong profile URL.

Reads from:
  results/comparison/Comparison_Ranked.csv   (produced by comparison.py)

Output:
  results/comparison/Outlier_Report.xlsx

Run via main.py (option [O]) or directly:
    python -m utils.outlier_report

Called automatically at the end of comparison.py when the comparison file exists.
"""

import sys
import pandas as pd
from pathlib import Path

if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent.parent))

import config

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── Thresholds ─────────────────────────────────────────────────────────────────
# Expected citation trend: Google Scholar > WoS > ScholarGPS.
#
# GS:WoS ratio typically ~2.5 (SD ~1.0–1.5).
#   Flag HIGH outliers (ratio > mean + SD_MULTIPLIER × SD) → WoS under-counting.
#   Low GS:WoS is not flagged — GS > WoS is the expected direction.
#
# SGPS:GS and SGPS:WoS ratios are expected to be < 1 (SGPS < GS, SGPS < WoS).
#   Flag LOW outliers (ratio < mean − SD_MULTIPLIER × SD) → SGPS under-counting.
#
# GS:WoS ratio fixed bounds (professor's domain knowledge: mean ~2.5, SD ~1.0).
# Flag if ratio < GS_WOS_LO or ratio > GS_WOS_HI — expect ~12 flagged faculty in a 50-person cohort.
GS_WOS_LO = 1.0
GS_WOS_HI = 4.0

# SD multiplier for ScholarGPS ratios (cohort-computed, less domain knowledge available).
SD_MULTIPLIER = 1.0

# Minimum GS citations needed to flag a zero-WoS case as anomalous.
# Lowered from 50 — faculty with even moderate GS activity should have some WoS.
MIN_GS_FOR_ZERO_WOS_FLAG = 20


def run() -> None:
    """
    Main entry point. Reads Comparison_Ranked.csv and writes Outlier_Report.xlsx.
    Called by comparison.py after it completes, or run directly via main.py [O].
    """
    print("\n[Outlier Report] Generating citation outlier report…")

    comp_csv = config.COMP_RESULTS / "Comparison_Ranked.csv"
    if not comp_csv.exists():
        print(f"  [ERROR] Comparison file not found: {comp_csv.name}")
        print("  Run the cross-source comparison first (option [C] in main.py).")
        return

    df = pd.read_csv(comp_csv)

    # ── Check which sources are available ─────────────────────────────────────
    has_wos  = "WoS_Total Citations"        in df.columns
    has_gs   = "Google_Total Citations"     in df.columns
    has_sgps = "ScholarGPS_Total Citations" in df.columns

    if not (has_wos and has_gs):
        print("  [ERROR] Comparison file must contain both WoS and Google Scholar results.")
        return

    # ── GS:WoS ratio analysis ─────────────────────────────────────────────────
    # GS:WoS = Google Scholar ÷ WoS. Expected trend: GS > WoS, so ratio > 1.
    # Typical cohort value is ~2.5 (SD ~1.0–1.5).
    # Flag HIGH outliers only (ratio > mean + SD_MULTIPLIER × SD):
    #   high ratio = WoS is significantly under-counting relative to GS.
    # Faculty with zero WoS are excluded from the ratio (flagged separately below).
    df_with_wos = df[df["WoS_Total Citations"] > 0].copy()
    df_with_wos["_gs_wos_ratio"] = (
        df_with_wos["Google_Total Citations"] / df_with_wos["WoS_Total Citations"]
    )

    gs_wos_vals = df_with_wos["_gs_wos_ratio"]
    print(f"  GS:WoS ratio  — median: {float(gs_wos_vals.median()):.2f}, "
          f"range: [{float(gs_wos_vals.min()):.2f}, {float(gs_wos_vals.max()):.2f}]")
    print(f"  Flag bounds   (fixed): ratio < {GS_WOS_LO} or ratio > {GS_WOS_HI}")

    # Build lookup: email → ratio
    gs_wos_lookup = (
        df_with_wos
        .set_index(config.COL_EMAIL)["_gs_wos_ratio"]
        .to_dict()
    )

    # ── ScholarGPS:GS ratio analysis ─────────────────────────────────────────
    # SGPS:GS = ScholarGPS ÷ GS. Expected: SGPS < GS, so ratio < 1.
    # Flag LOW outliers (ratio < mean − SD_MULTIPLIER × SD):
    #   low ratio = ScholarGPS is significantly under-counting relative to GS,
    #   likely due to a wrong or mismatched ScholarGPS profile URL.
    sgps_lo = None
    sgps_ratio_lookup = {}
    if has_sgps:
        df_with_gs = df[df["Google_Total Citations"] > 0].copy()
        df_with_gs["_sgps_gs_ratio"] = (
            df_with_gs["ScholarGPS_Total Citations"] / df_with_gs["Google_Total Citations"]
        )
        sgps_vals = df_with_gs["_sgps_gs_ratio"]
        sgps_mean = float(sgps_vals.mean())
        sgps_sd   = float(sgps_vals.std())
        sgps_lo   = sgps_mean - SD_MULTIPLIER * sgps_sd
        sgps_ratio_lookup = (
            df_with_gs
            .set_index(config.COL_EMAIL)["_sgps_gs_ratio"]
            .to_dict()
        )
        print(f"  SGPS:GS ratio  — mean: {sgps_mean:.2f}, SD: {sgps_sd:.2f}, "
              f"range: [{float(sgps_vals.min()):.2f}, {float(sgps_vals.max()):.2f}]")
        print(f"  Low threshold  (mean − {SD_MULTIPLIER}×SD): {sgps_lo:.2f}")

    # ── Build per-row flags ────────────────────────────────────────────────────
    gs_wos_ratio_col  = []
    sgps_gs_ratio_col = []
    flag_col          = []

    for _, row in df.iterrows():
        email  = row.get(config.COL_EMAIL, "")
        gs     = row["Google_Total Citations"]
        wos    = row["WoS_Total Citations"]
        flags  = []

        # --- Zero WoS with significant GS ---
        if wos == 0 and gs >= MIN_GS_FOR_ZERO_WOS_FLAG:
            flags.append(f"Zero WoS citations (GS = {int(gs)}) — verify WoS ResearchID")

        # --- GS:WoS ratio (either direction = anomalous) ---
        # High ratio: WoS under-counting relative to GS → verify ResearchID
        # Low ratio:  GS may be over-counting (wrong/merged GS profile) → verify GS profile
        ratio = gs_wos_lookup.get(email)
        gs_wos_ratio_col.append(round(ratio, 2) if ratio is not None else None)
        if ratio is not None and ratio > GS_WOS_HI:
            flags.append(f"High GS:WoS ratio ({ratio:.2f} > {GS_WOS_HI}) — WoS may be undercounting; verify ResearchID")
        elif ratio is not None and ratio < GS_WOS_LO:
            flags.append(f"Low GS:WoS ratio ({ratio:.2f} < {GS_WOS_LO}) — GS may be overcounting; verify Google Scholar profile")

        # --- SGPS:GS ratio (low = ScholarGPS under-counting / wrong profile) ---
        sgps_ratio = sgps_ratio_lookup.get(email) if has_sgps else None
        sgps_gs_ratio_col.append(round(sgps_ratio, 2) if sgps_ratio is not None else None)
        if sgps_ratio is not None and sgps_lo is not None and sgps_ratio < sgps_lo:
            flags.append(f"Low SGPS:GS ratio ({sgps_ratio:.2f} < {sgps_lo:.2f}) — verify ScholarGPS profile URL")

        flag_col.append(" | ".join(flags))

    df["GS:WoS Ratio"]  = gs_wos_ratio_col
    if has_sgps:
        df["SGPS:GS Ratio"] = sgps_gs_ratio_col
    df["Outlier Flags"] = flag_col

    # ── Filter to flagged rows only ────────────────────────────────────────────
    df_flagged = df[df["Outlier Flags"] != ""].copy().reset_index(drop=True)

    print(f"\n  Faculty checked   : {len(df)}")
    print(f"  Outliers flagged  : {len(df_flagged)}")

    if df_flagged.empty:
        print("  No outliers found.")
        return

    print("\n  Flagged faculty:")
    for _, row in df_flagged.iterrows():
        name = f"{row[config.COL_FIRST_NAME]} {row[config.COL_LAST_NAME]}"
        print(f"    {name:<30} {row['Outlier Flags']}")

    # ── Build output DataFrame ─────────────────────────────────────────────────
    priority = (
        config.FACULTY_KEY_COLS
        + ["Google_Total Citations", "WoS_Total Citations"]
        + (["ScholarGPS_Total Citations"] if has_sgps else [])
        + ["GS:WoS Ratio"]
        + (["SGPS:GS Ratio"] if has_sgps else [])
        + ["Outlier Flags"]
    )
    available  = [c for c in priority if c in df_flagged.columns]
    remaining  = [c for c in df_flagged.columns if c not in available]
    df_out     = df_flagged[available + remaining]

    # ── Save Excel ─────────────────────────────────────────────────────────────
    config.COMP_RESULTS.mkdir(parents=True, exist_ok=True)
    out_path = config.COMP_RESULTS / "Outlier_Report.xlsx"
    df_out.to_excel(out_path, index=False)

    if OPENPYXL_OK:
        _format_excel(out_path)

    print(f"\n[Outlier Report] Saved → {out_path.name}")
    print(f"  Review flagged rows and verify WoS IDs / ScholarGPS profile URLs as needed.")


def _format_excel(path: Path) -> None:
    """Apply colour formatting to the outlier report Excel file."""
    wb = load_workbook(path)
    ws = wb.active

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark blue
    ZERO_WOS    = PatternFill("solid", fgColor="FF6B6B")   # red  — zero WoS
    RATIO_FLAG  = PatternFill("solid", fgColor="FFD700")   # gold — ratio outlier
    SGPS_FLAG   = PatternFill("solid", fgColor="FFA500")   # orange — ScholarGPS outlier
    BORDER = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    header_vals   = [cell.value for cell in ws[1]]
    flags_col_idx = header_vals.index("Outlier Flags")   # 0-based

    # Header row
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER

    # Data rows — colour by primary flag type
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        flag_text = str(row[flags_col_idx].value or "")

        if "Zero WoS" in flag_text:
            fill = ZERO_WOS
        elif "SGPS" in flag_text:
            fill = SGPS_FLAG
        else:
            fill = RATIO_FLAG

        for cell in row:
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Column widths
    col_widths = {
        "Last Name": 16,  "First Name": 14,  "Department": 12,
        "Email": 32,      "Faculty Type": 12,
        "Google_Total Citations": 14,  "WoS_Total Citations": 14,
        "ScholarGPS_Total Citations": 14,
        "GS:WoS Ratio": 12,  "SGPS:GS Ratio": 12,
        "Outlier Flags": 65,
    }
    for i, col_name in enumerate(header_vals, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    run()
