"""
update_faculty_list.py
----------------------
Merges the new semester's faculty list with researcher IDs from the prior year.

Run this at the start of each year before running the data collection pipeline.

HOW TO USE EACH YEAR
--------------------
1. Update the three file path constants below (NEW_LIST, PRIOR_LIST, OUT_PATH).
2. Run:  python update_faculty_list.py
3. Open the output Excel file and review the "Review Required" sheet:
   - UNSURE rows: check whether it's the same person with a different name format,
     or a genuinely new person.
   - NEW rows: contact these faculty for their WoS / Google Scholar IDs.
4. Fill in any missing IDs and save the file as your new master faculty list
   (e.g., 2027_Spring_Faculty_List.xlsx).

MATCHING LOGIC
--------------
  MATCHED  — email match, OR exact last+first name match
             → IDs carried over automatically, no action needed

  UNSURE   — same last name + same first initial, but no email/exact name match
             → Could be same person (nickname, typo, name change) or different person
             → Listed in "Review Required" sheet for manual check

  NEW      — no match at all
             → Listed in "Review Required" sheet; contact for IDs

OUTPUT
------
  Sheet 1 "Updated Faculty List"  — full merged list, colour-coded by status
  Sheet 2 "Review Required"       — only UNSURE + NEW rows, with notes
"""

import unicodedata
from difflib import SequenceMatcher
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── !! UPDATE THESE EACH YEAR !! ─────────────────────────────────────────────
NEW_LIST   = "../Marshall Faculty List_February 2026.xlsx"  # new semester's raw list
PRIOR_LIST = "../2025_Fall_Faculty_List.xlsx"               # prior year's list with IDs
OUT_PATH   = "../2026_Spring_Faculty_List.xlsx"             # output file name

# Fuzzy match threshold (0.0–1.0).
# Two names scoring at or above this are flagged UNSURE for human review.
# Raise to catch fewer (stricter); lower to catch more (broader).
#   0.85 = catches typos, hyphens, accents, nicknames; misses very different names
#   0.75 = broader — also catches short vs long first names ("Jo" vs "Joanne")
FUZZY_THRESHOLD = 0.82
# ─────────────────────────────────────────────────────────────────────────────

# ID columns to carry over from the prior year list
ID_COLS = [
    "Google Scholar Profile Link", "ORCID", "WoS", "SCOPUS_ID",
    "Middle initials", "Alternative first name",
    "Alternative surname/ family name", "ORCID Certainty",
]

# Columns from the new list to keep (in order)
BASE_COLS = [
    "Last Name", "First Name", "Email",
    "Marshall Dept Abbreviation", "Rank", "Faculty Type", "Faculty Status",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _norm(series: pd.Series) -> pd.Series:
    """Lowercase + strip for reliable string matching."""
    return series.fillna("").astype(str).str.strip().str.lower()


def _normalize_for_fuzzy(text: str) -> str:
    """
    Prepare a name string for fuzzy comparison:
    - Remove accents (José → jose, Müller → muller)
    - Lowercase and strip
    - Collapse hyphens/dashes to spaces (Garcia-Lopez → garcia lopez)
    """
    # Strip accents via Unicode decomposition
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_text = nfkd.encode("ascii", "ignore").decode("ascii")
    return ascii_text.lower().strip().replace("-", " ").replace("  ", " ")


def _similarity(name_a: str, name_b: str) -> float:
    """Return 0.0–1.0 similarity ratio between two normalized name strings."""
    a = _normalize_for_fuzzy(name_a)
    b = _normalize_for_fuzzy(name_b)
    return SequenceMatcher(None, a, b).ratio()


# ── 1. Load data ──────────────────────────────────────────────────────────────
new_df   = pd.read_excel(NEW_LIST)
prior_df = pd.read_excel(PRIOR_LIST)

# Normalise matching keys
for df in (new_df, prior_df):
    df["_email"] = _norm(df["Email"])
    df["_last"]  = _norm(df["Last Name"])
    df["_first"] = _norm(df["First Name"])

# ── 2. Build prior-year lookup tables ────────────────────────────────────────
# Available ID columns (only keep those that actually exist in the prior list)
available_id_cols = [c for c in ID_COLS if c in prior_df.columns]

prior_dedup_email = prior_df.drop_duplicates("_email", keep="first")
prior_dedup_name  = prior_df.drop_duplicates(["_last", "_first"], keep="first")

by_email = prior_dedup_email.set_index("_email")
by_name  = prior_dedup_name.set_index(["_last", "_first"])

# Flat list of prior records for full-scan fuzzy matching
prior_records = prior_df.to_dict("records")

# ── 3. Match each new faculty member ─────────────────────────────────────────
statuses    = []   # "MATCHED" | "UNSURE" | "NEW"
id_rows     = []   # ID column values
fuzzy_notes = []   # Human-readable note for UNSURE rows

for _, row in new_df.iterrows():
    email    = row["_email"]
    last     = row["_last"]
    first    = row["_first"]
    new_full = f"{first} {last}"

    # --- Email match (most reliable) ---
    if email and email in by_email.index:
        ids = by_email.loc[email, available_id_cols].to_dict()
        statuses.append("MATCHED")
        id_rows.append(ids)
        fuzzy_notes.append("")
        continue

    # --- Exact name match ---
    if (last, first) in by_name.index:
        ids = by_name.loc[(last, first), available_id_cols].to_dict()
        statuses.append("MATCHED")
        id_rows.append(ids)
        fuzzy_notes.append("")
        continue

    # --- Fuzzy name match (similarity score across ALL prior faculty) ---
    # Searches the full prior list so it catches name changes, hyphenation,
    # accents, typos, and nickname variations regardless of last name.
    best_score  = 0.0
    best_record = None
    for prior in prior_records:
        prior_full = f"{prior['_first']} {prior['_last']}"
        score = _similarity(new_full, prior_full)
        if score > best_score:
            best_score  = score
            best_record = prior

    if best_score >= FUZZY_THRESHOLD:
        ids = {c: best_record.get(c) for c in available_id_cols}
        statuses.append("UNSURE")
        id_rows.append(ids)
        prior_name  = f"{best_record['First Name']} {best_record['Last Name']}"
        prior_email = best_record.get("Email", "")
        fuzzy_notes.append(
            f"Prior year: {prior_name} ({prior_email})"
            f" — similarity {best_score:.0%} — verify same person"
        )
        continue

    # --- No match ---
    statuses.append("NEW")
    id_rows.append({c: None for c in available_id_cols})
    fuzzy_notes.append("Contact for WoS / Google Scholar / ORCID IDs")

# ── 4. Build merged DataFrame ────────────────────────────────────────────────
id_data = pd.DataFrame(id_rows, index=new_df.index)

# Keep only base cols that exist in the new list
existing_base = [c for c in BASE_COLS if c in new_df.columns]
merged = pd.concat(
    [new_df[existing_base].reset_index(drop=True),
     id_data.reset_index(drop=True)],
    axis=1,
)
merged["Status"]  = statuses
merged["Notes"]   = fuzzy_notes

# ── 5. Print summary ──────────────────────────────────────────────────────────
n_matched = statuses.count("MATCHED")
n_unsure  = statuses.count("UNSURE")
n_new     = statuses.count("NEW")
total     = len(merged)

print(f"Total faculty in new list : {total}")
print(f"  MATCHED (IDs carried over)  : {n_matched}")
print(f"  UNSURE  (check manually)    : {n_unsure}")
print(f"  NEW     (contact for IDs)   : {n_new}")

if n_unsure:
    print("\nUNSURE — verify these are the same person:")
    unsure_df = merged[merged["Status"] == "UNSURE"][
        ["Last Name", "First Name", "Email", "Notes"]
    ]
    print(unsure_df.to_string(index=False))

if n_new:
    print("\nNEW faculty — contact for researcher IDs:")
    new_fac = merged[merged["Status"] == "NEW"][
        ["Last Name", "First Name", "Email",
         existing_base[3] if len(existing_base) > 3 else "Last Name"]
    ]
    print(new_fac.to_string(index=False))

# ── 6. Write Sheet 1: full updated list ──────────────────────────────────────
col_order = existing_base + available_id_cols + ["Status", "Notes"]
merged = merged[col_order]

with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    merged.to_excel(writer, sheet_name="Updated Faculty List", index=False)

    # Sheet 2: Review Required (UNSURE + NEW only)
    review_df = merged[merged["Status"].isin(["UNSURE", "NEW"])].copy()
    review_df.to_excel(writer, sheet_name="Review Required", index=False)

# ── 7. Format both sheets ─────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
ID_FILL      = PatternFill("solid", fgColor="D9E8F5")   # light blue — ID columns
NEW_FILL     = PatternFill("solid", fgColor="FFFF99")   # yellow — new faculty
UNSURE_FILL  = PatternFill("solid", fgColor="FFD966")   # amber — unsure
MATCHED_FILL = PatternFill("solid", fgColor="E2EFDA")   # light green — matched
STATUS_NEW_FILL    = PatternFill("solid", fgColor="FF9999")   # red — status cell
STATUS_UNSURE_FILL = PatternFill("solid", fgColor="F4B942")   # orange — status cell
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COL_WIDTHS = {
    "Last Name": 18, "First Name": 16, "Email": 34,
    "Marshall Dept Abbreviation": 12, "Rank": 10,
    "Faculty Type": 12, "Faculty Status": 12,
    "WoS": 18, "Google Scholar Profile Link": 50,
    "ORCID": 24, "ORCID Certainty": 12, "SCOPUS_ID": 16,
    "Middle initials": 10, "Alternative first name": 18,
    "Alternative surname/ family name": 26,
    "Status": 12, "Notes": 45,
}


def _format_sheet(ws, id_col_names: set[str]) -> None:
    header_vals = [cell.value for cell in ws[1]]
    status_idx  = header_vals.index("Status") + 1 if "Status" in header_vals else None

    # Header
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER

    # Data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status = str(row[status_idx - 1].value) if status_idx else ""
        row_fill = (
            UNSURE_FILL  if status == "UNSURE"  else
            NEW_FILL     if status == "NEW"      else
            MATCHED_FILL
        )
        for cell in row:
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            col_name = header_vals[cell.column - 1]

            if col_name == "Status":
                cell.fill = STATUS_UNSURE_FILL if status == "UNSURE" else \
                            STATUS_NEW_FILL    if status == "NEW"     else \
                            MATCHED_FILL
                cell.font = Font(bold=True)
            elif col_name in id_col_names:
                cell.fill = ID_FILL
            elif status != "MATCHED":
                cell.fill = row_fill

    # Column widths
    for i, col_name in enumerate(header_vals, start=1):
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(col_name, 14)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


wb = load_workbook(OUT_PATH)
id_col_set = set(available_id_cols)

for sheet_name in ["Updated Faculty List", "Review Required"]:
    if sheet_name in wb.sheetnames:
        _format_sheet(wb[sheet_name], id_col_set)

wb.save(OUT_PATH)
print(f"\nSaved: {OUT_PATH}")
print(f"  Sheet 1 'Updated Faculty List' — {total} faculty")
print(f"  Sheet 2 'Review Required'      — {n_unsure + n_new} rows to check")
